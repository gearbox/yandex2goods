from pathlib import Path
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import datetime


# Indent XML into a beautified format
def indent(elem, level=0):
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


def verify_input(text: str):
    return text.replace('"', '&quot;').replace('&', '&amp;').replace('>', '&gt;').\
        replace('<', '&lt;').replace('\'', '&apos;')


def build_tree(file_name, worksheet, categories: dict, max_rows_limit=None):
    max_row = None
    if 0 < max_rows_limit < worksheet.max_row:
        max_row = max_rows_limit + 1
    catalogue = ET.Element('yml_catalog', date=str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M')))
    shop = ET.SubElement(catalogue, 'shop')
    ET.SubElement(shop, 'name').text = 'Vitavim'
    ET.SubElement(shop, 'company').text = 'ООО "Жизнь"'
    ET.SubElement(shop, 'url').text = 'https://vitavim.ru/'
    shop_currencies = ET.SubElement(shop, 'currencies')
    ET.SubElement(shop_currencies, 'currency', id='RUR', rate='1')
    categories_section = ET.SubElement(shop, 'categories')
    for cat_name, cat_id in categories.items():
        ET.SubElement(categories_section, 'category', id=cat_id).text = cat_name
    # shipment_options = ET.SubElement(shop, 'shipment-options')
    offers = ET.SubElement(shop, "offers")

    for product in worksheet.iter_rows(min_row=2, max_row=max_row):
        available = 'true' if product[1].value == 'В наличии' or product[1].value is None else 'false'
        offer = ET.SubElement(offers, 'offer', id=str(product[0].value), available=available)
        ET.SubElement(offer, 'url').text = product[7].value
        ET.SubElement(offer, 'name').text = verify_input(product[9].value)
        ET.SubElement(offer, 'price').text = verify_input(str(product[11].value))
        ET.SubElement(offer, 'categoryId').text = categories[product[10].value]
        ET.SubElement(offer, 'picture').text = product[14].value
        ET.SubElement(offer, 'vendor').text = verify_input(product[8].value)
        ET.SubElement(offer, 'description').text = verify_input(product[15].value)
    indent(catalogue)

    tree = ET.ElementTree(catalogue)
    tree.write(file_name, xml_declaration=True, encoding='utf-8', method="xml")


def read_files(path: Path, file_types: tuple):
    if path.is_dir():
        input_files = [item for item in path.iterdir() if item.name.endswith(file_types)]
        return input_files
    return False


if __name__ == "__main__":
    # Settings
    # Список категорий в формате {'name1': 'id1', 'name2': 'id2', etc}
    product_categories = {'Суперфуды': '1', 'Масло растительное': '2'}

    # Количество товаров конвертируемых в xml, 0 - без ограничения
    products_limit = 0

    project_dir_path = Path(__file__).parent
    project_input_dir = project_dir_path / 'in'
    project_results_dir = project_dir_path / 'out'
    Path(project_results_dir).mkdir(exist_ok=True)
    input_xls = project_input_dir / 'goods.xlsx'

    types = ('.xls', '.xlsx')
    files_grabbed = read_files(project_input_dir, types)

    for f in files_grabbed:
        output_xml = project_results_dir / (str(datetime.datetime.now().date()) + '_' + f.name.split('.')[0] + '.xml')
        wb = load_workbook(f)
        for sheet in wb.worksheets:
            header = list(sheet.rows)[0]
            if sheet.max_row > 1:
                build_tree(output_xml, sheet, product_categories, max_rows_limit=products_limit)
